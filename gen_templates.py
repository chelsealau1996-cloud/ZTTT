#!/usr/bin/env python3
"""Generate 发票模板导入格式示例.xlsx from 形式发票模板合并 (1).xlsx.

Each sheet = one template. RED=field key, BLACK=static text.
Scoped field keys (行项目.描述.US) where rules differ across templates.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── style constants ──────────────────────────────────────────────────
black_font  = Font(name="Arial", size=9, color="000000")
red_font    = Font(name="Consolas", size=9, color="FF0000")
title_font  = Font(name="Arial", size=14, bold=True, color="000000")
jp_title    = Font(name="Arial", size=11, bold=True, color="000000")
header_fill = PatternFill("solid", fgColor="4472C4")
sub_fill    = PatternFill("solid", fgColor="D9E2F3")
gray_font   = Font(name="Arial", size=9, color="666666")
bank_font   = Font(name="Microsoft YaHei", size=7, bold=True, color="000000")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

M = Side(style='medium')
T = Side(style='thin')
N = None

M_ALL   = Border(left=M, right=M, top=M, bottom=M)
M_LR    = Border(left=M, right=M)
M_LT    = Border(left=M, top=M)
M_LB    = Border(left=M, bottom=M)
M_LRT   = Border(left=M, right=M, top=M)
M_LRB   = Border(left=M, right=M, bottom=M)
M_LRTB  = Border(left=M, right=M, top=M, bottom=M)
M_L     = Border(left=M)
M_B     = Border(bottom=M)
M_RTB   = Border(right=M, top=M, bottom=M)
M_LTB   = Border(left=M, top=M, bottom=M)
M_R     = Border(right=M)
M_RT    = Border(right=M, top=M)
M_RB    = Border(right=M, bottom=M)
M_T     = Border(top=M)
M_TB    = Border(top=M, bottom=M)

T_ALL   = Border(left=T, right=T, top=T, bottom=T)
T_LR    = Border(left=T, right=T)
T_LRB   = Border(left=T, right=T, bottom=T)
T_LRT   = Border(left=T, right=T, top=T)
T_LTB   = Border(left=T, top=T, bottom=T)
T_RTB   = Border(right=T, top=T, bottom=T)
T_LR_TB = Border(left=T, right=T, top=T, bottom=T)
T_B     = Border(bottom=T)
T_TB    = Border(top=T, bottom=T)
T_L     = Border(left=T)
T_R     = Border(right=T)
T_T     = Border(top=T)
T_LT    = Border(left=T, top=T)
T_RT    = Border(right=T, top=T)
T_LB    = Border(left=T, bottom=T)
T_RB    = Border(right=T, bottom=T)

def sc(ws, r, c, v=None, fnt=None, fill=None, al=None, border=None):
    cell = ws.cell(row=r, column=c, value=v)
    if fnt: cell.font = fnt
    if fill: cell.fill = fill
    if al: cell.alignment = al
    if border: cell.border = border
    return cell

def set_cols(ws, widths):
    for k, w in widths.items():
        ws.column_dimensions[k].width = w

# ══════════════════════════════════════════════════════════════════════
#  1-4: Canada (GST 5% / HST 13%  x  positive / credit note)
# ══════════════════════════════════════════════════════════════════════
CA_TAX_INFO = {
    'GST5':  ('GST 5%',    'GST/HST NO.: 772438263 RT0002'),
    'HST13': ('HST 13%',   'GST/HST NO.: 772438263 RT0001'),
}

def build_ca(ws, tax_label, tax_num, is_credit=False):
    set_cols(ws, {'B': 34, 'C': 18, 'D': 18, 'E': 20})
    inv_title = "Credit Note" if is_credit else "INVOICE"
    inv_label = "Credit. No." if is_credit else "Invoice. No."

    sc(ws, 1, 2, "GOODCANG LOGISTICS CANADA CORP.", fnt=jp_title,
       al=Alignment(horizontal='center'))
    sc(ws, 2, 2, "5 Paget Road, Brampton, Ontario, L6T5S2", fnt=black_font,
       al=Alignment(horizontal='center'))
    sc(ws, 3, 2, tax_num, fnt=black_font,
       al=Alignment(horizontal='center'))
    sc(ws, 4, 2, "PST NO.: PST-1472-6656", fnt=black_font,
       al=Alignment(horizontal='center'))
    ws.merge_cells('B1:E1')
    ws.merge_cells('B2:E2')
    ws.merge_cells('B3:E3')
    ws.merge_cells('B4:E4')

    sc(ws, 6, 2, "Bill to:", fnt=black_font)
    sc(ws, 7, 2, "客户名称", fnt=red_font)
    sc(ws, 8, 2, "客户地址", fnt=red_font)

    sc(ws, 10, 3, inv_title, fnt=Font(name="Arial", size=16, bold=True, color="000000"),
       al=Alignment(horizontal='center', vertical='center'))
    ws.merge_cells('C10:E11')

    sc(ws, 12, 3, inv_label, fnt=black_font, border=T_ALL)
    sc(ws, 12, 4, "Invoice Date", fnt=black_font, border=T_ALL)
    sc(ws, 12, 5, "Term", fnt=black_font, border=T_ALL)
    neg = ".NEG" if is_credit else ""

    sc(ws, 13, 3, "发票号码.CA", fnt=red_font, border=T_ALL)
    sc(ws, 13, 4, "开票日期.US", fnt=red_font, border=T_ALL)
    sc(ws, 13, 5, "付款周期", fnt=red_font, border=T_ALL)

    # Column headers
    sc(ws, 16, 2, "Description", fnt=black_font, al=Alignment(horizontal='center'),
       border=Border(left=T, right=T, top=T, bottom=M))
    sc(ws, 16, 3, "Amount", fnt=black_font, al=Alignment(horizontal='center'),
       border=Border(left=T, right=T, top=T, bottom=M))
    sc(ws, 16, 4, tax_label, fnt=black_font, al=Alignment(horizontal='center'),
       border=Border(left=T, right=T, top=T, bottom=M))
    sc(ws, 16, 5, "Total", fnt=black_font, al=Alignment(horizontal='center'),
       border=Border(left=T, right=T, top=T, bottom=M))

    # Line items (rows 17-25)
    for r in range(17, 26):
        for c in range(2, 6):
            sc(ws, r, c, None, border=T_LRB)
    sc(ws, 17, 2, "行项目.描述.US", fnt=red_font, border=T_LRB)
    sc(ws, 17, 3, f"行项目.金额{neg}", fnt=red_font, border=T_LRB)
    sc(ws, 17, 4, f"行项目.税额{neg}", fnt=red_font, border=T_LRB)
    sc(ws, 17, 5, f"行项目.含税金额{neg}", fnt=red_font, border=T_LRB)

    # Total
    for c in range(2, 6):
        sc(ws, 26, c, None, border=T_ALL)
    sc(ws, 26, 2, "TOTAL", fnt=black_font, border=T_ALL)
    sc(ws, 26, 3, f"合计金额{neg}", fnt=red_font, border=T_ALL)
    sc(ws, 26, 4, f"合计税额{neg}", fnt=red_font, border=T_ALL)
    sc(ws, 26, 5, f"合计含税金额{neg}", fnt=red_font, border=T_ALL)

    sc(ws, 27, 2, "发票备注", fnt=red_font)

    # Bank info (static)
    sc(ws, 28, 2, "Please remit to :", fnt=black_font)
    sc(ws, 29, 2, "Beneficiary: GOODCANG LOGISTICS CANADA CORP.", fnt=black_font)
    sc(ws, 30, 2, "Beneficiary's address: 5 Paget Road, Brampton, ON L6T 5S2", fnt=black_font)
    sc(ws, 31, 2, "Bank name: Aberdeen Centre Branch", fnt=black_font)
    sc(ws, 32, 2, "Bank Address: 1160-4151 Hazelbridge Way, Richmond V6X 4J7", fnt=black_font)
    sc(ws, 33, 2, "Bank Transit: 03395", fnt=black_font)
    sc(ws, 34, 2, "Fiancial Institution: 003", fnt=black_font)
    sc(ws, 35, 2, "SWIFT Code: ROYCCAT2", fnt=black_font)
    sc(ws, 36, 2, "Account#: 033951001429(CAD)", fnt=black_font)

# ══════════════════════════════════════════════════════════════════════
#  5-8: Vancouver (GST 0%/5% / HST 13-15%  x  positive / credit note)
# ══════════════════════════════════════════════════════════════════════
VC_TAX_INFO = {
    'GST':  ('GST 0%',  'GST/HST NO.: 794845438 RT0002'),
    'HST':  ('HST 13%', 'GST/HST NO.: 794845438 RT0001'),
}

def build_vc(ws, tax_label, tax_num, is_credit=False):
    set_cols(ws, {'B': 34, 'C': 18, 'D': 18, 'E': 20})
    inv_title = "Credit Note" if is_credit else "INVOICE"
    inv_label = "Credit. No." if is_credit else "Invoice. No."

    sc(ws, 1, 2, "Zenda Logistics Corp.", fnt=jp_title,
       al=Alignment(horizontal='center'))
    sc(ws, 2, 2, "105-19159 22ND AVENUE SURREY BC V3Z 3S6 CANADA", fnt=black_font,
       al=Alignment(horizontal='center'))
    sc(ws, 3, 2, tax_num, fnt=black_font,
       al=Alignment(horizontal='center'))
    ws.merge_cells('B1:E1')
    ws.merge_cells('B2:E2')
    ws.merge_cells('B3:E3')

    sc(ws, 6, 2, "Bill to:", fnt=black_font)
    sc(ws, 7, 2, "客户名称", fnt=red_font)
    sc(ws, 8, 2, "客户地址", fnt=red_font)

    sc(ws, 10, 3, inv_title, fnt=Font(name="Arial", size=16, bold=True, color="000000"),
       al=Alignment(horizontal='center', vertical='center'))
    ws.merge_cells('C10:E11')

    sc(ws, 12, 3, inv_label, fnt=black_font, border=T_ALL)
    sc(ws, 12, 4, "Invoice Date", fnt=black_font, border=T_ALL)
    sc(ws, 12, 5, "Term", fnt=black_font, border=T_ALL)
    neg = ".NEG" if is_credit else ""

    sc(ws, 13, 3, "发票号码.VC", fnt=red_font, border=T_ALL)
    sc(ws, 13, 4, "开票日期.VC", fnt=red_font, border=T_ALL)
    sc(ws, 13, 5, "付款周期", fnt=red_font, border=T_ALL)

    sc(ws, 15, 2, "CURRENCY: CANADIAN DOLLAR", fnt=black_font)

    # Column headers
    sc(ws, 16, 2, "Description", fnt=black_font, al=Alignment(horizontal='center'),
       border=Border(left=T, right=T, top=T, bottom=M))
    sc(ws, 16, 3, "Amount", fnt=black_font, al=Alignment(horizontal='center'),
       border=Border(left=T, right=T, top=T, bottom=M))
    sc(ws, 16, 4, tax_label, fnt=black_font, al=Alignment(horizontal='center'),
       border=Border(left=T, right=T, top=T, bottom=M))
    sc(ws, 16, 5, "Total", fnt=black_font, al=Alignment(horizontal='center'),
       border=Border(left=T, right=T, top=T, bottom=M))

    # Line items (rows 17-27)
    for r in range(17, 28):
        for c in range(2, 6):
            sc(ws, r, c, None, border=T_LRB)
    sc(ws, 17, 2, "行项目.描述.US", fnt=red_font, border=T_LRB)
    sc(ws, 17, 3, f"行项目.金额{neg}", fnt=red_font, border=T_LRB)
    sc(ws, 17, 4, f"行项目.税额{neg}", fnt=red_font, border=T_LRB)
    sc(ws, 17, 5, f"行项目.含税金额{neg}", fnt=red_font, border=T_LRB)

    # Total
    for c in range(2, 6):
        sc(ws, 28, c, None, border=T_ALL)
    sc(ws, 28, 2, "TOTAL", fnt=black_font, border=T_ALL)
    sc(ws, 28, 3, f"合计金额{neg}", fnt=red_font, border=T_ALL)
    sc(ws, 28, 4, f"合计税额{neg}", fnt=red_font, border=T_ALL)
    sc(ws, 28, 5, f"合计含税金额{neg}", fnt=red_font, border=T_ALL)

    sc(ws, 29, 2, "发票备注", fnt=red_font)

    # Bank info
    sc(ws, 30, 2, "Please remit to :", fnt=black_font)
    sc(ws, 31, 2, "Beneficiary: Zenda Logistics Corp.", fnt=black_font)
    sc(ws, 31, 4, "Beneficiary's address: 105-19159 22ND AVENUE SURREY BC V3Z 3S6 CANADA", fnt=black_font)
    sc(ws, 32, 2, "Bank name: Scotiabank", fnt=black_font)
    sc(ws, 32, 4, "SWIFT Code: NOSCCATT", fnt=black_font)
    sc(ws, 33, 2, "Account#: 0114618", fnt=black_font)

# ══════════════════════════════════════════════════════════════════════
#  9-12: US-EL / US-GA (positive / credit note)
# ══════════════════════════════════════════════════════════════════════
US_EL_BANK = [
    "Beneficiary Name: (required)：Us Elogistics Service Corp",
    "Beneficiary Account Number: (required)：8003139485",
    "Bank Routing Number: (domestic wires)：3 | 2 | 2 | 0 | 7 | 0 | 3 | 8 | 1",
    "International：B|K| | |U|S| |3| |C| | |N| |A| |",
    "Name of Bank: Citibank New York",
    "Bank SWIFT Address: CITIUS33",
    "Account: 8003139485",
]
US_GA_BANK = [
    "Beneficiary Name: (required)：ELogistics (GA) Service Corp.",
    "Beneficiary Account Number: (required)：8003135707",
    "Bank Routing Number: (domestic wires)：0 | 6 | 1 | 1 | 0 | 3 | 8 | 8 | 4",
    "International：B|K| | |U|S| |3| |C| | | |N| |A| |",
    "Name of Bank: Citibank New York",
    "Bank SWIFT Address: CITIUS33",
    "Account: 8003135707",
]

def build_el(ws, company_lines, is_credit=False):
    """US-EL positive or credit note. company_lines=[name, addr1, addr2]."""
    set_cols(ws, {'A': 36, 'B': 14, 'C': 8, 'D': 14, 'E': 16})

    if is_credit:
        # Credit note layout
        sc(ws, 1, 1, company_lines[0], fnt=jp_title,
           al=Alignment(horizontal='left', vertical='center', wrap_text=True))
        for i, line in enumerate(company_lines[1:], 2):
            sc(ws, i, 1, line.strip(), fnt=black_font,
               al=Alignment(horizontal='left', vertical='center', wrap_text=True)) if line.strip() else None
        ws.merge_cells('A1:A3')
        sc(ws, 1, 4, "Credit Note", fnt=title_font, al=Alignment(horizontal='right'))
        ws.merge_cells('D1:E1')
        sc(ws, 2, 4, "Date", fnt=black_font, al=Alignment(horizontal='right'))
        sc(ws, 2, 5, "credit note #", fnt=black_font, al=Alignment(horizontal='right'))
        sc(ws, 3, 4, "开票日期.US", fnt=red_font, al=Alignment(horizontal='right'))
        sc(ws, 3, 5, "发票号码.EL", fnt=red_font, al=Alignment(horizontal='right'))

        ws.merge_cells('D5:E7')
        sc(ws, 5, 1, "     Bill  To", fnt=black_font)
        sc(ws, 6, 1, "客户名称", fnt=red_font)
        sc(ws, 7, 1, "客户地址", fnt=red_font)

        # Column headers
        for c, v in {1:"Description", 3:"Qty", 4:"Rate", 5:"Amount"}.items():
            sc(ws, 9, c, v, fnt=black_font, al=Alignment(horizontal='center'))
        ws.merge_cells('A9:B9')
        # Line item template
        for c, v in {1:"行项目.描述.US", 3:"行项目.数量", 4:"行项目.单价.NEG", 5:"行项目.金额.NEG"}.items():
            sc(ws, 10, c, v, fnt=red_font, al=Alignment(wrap_text=True, vertical='top'))
        ws.merge_cells('A10:B10')
        # Total + Memo
        sc(ws, 11, 1, "发票备注", fnt=red_font, al=Alignment(horizontal='left', vertical='center'))
        ws.merge_cells('A11:C11')
        sc(ws, 11, 4, "Total", fnt=black_font)
        sc(ws, 11, 5, "合计金额.NEG", fnt=red_font)
        # Bank
        for i, line in enumerate(US_EL_BANK):
            sc(ws, 13+i, 1, line, fnt=bank_font)
        sc(ws, 13, 3, "开票主体", fnt=red_font, al=Alignment(vertical='top'))
        sc(ws, 14, 3, "收款信息", fnt=red_font, al=Alignment(vertical='top'))
        sc(ws, 15, 3, None)
    else:
        # Positive invoice layout
        sc(ws, 1, 1, company_lines[0], fnt=jp_title,
           al=Alignment(horizontal='left', vertical='center', wrap_text=True))
        for i, line in enumerate(company_lines[1:], 2):
            sc(ws, i, 1, line.strip(), fnt=black_font,
               al=Alignment(horizontal='left', vertical='center', wrap_text=True)) if line.strip() else None
        ws.merge_cells('A1:A3')
        ws.merge_cells('B1:D1')
        sc(ws, 1, 5, "Invoice", fnt=title_font, al=Alignment(horizontal='right'))
        sc(ws, 2, 4, "Date", fnt=black_font, al=Alignment(horizontal='right'))
        sc(ws, 2, 5, "Invoice #", fnt=black_font, al=Alignment(horizontal='right'))
        sc(ws, 3, 4, "开票日期.US", fnt=red_font, al=Alignment(horizontal='right'))
        sc(ws, 3, 5, "发票号码.EL", fnt=red_font, al=Alignment(horizontal='right'))

        ws.merge_cells('B5:C9')
        ws.merge_cells('A7:A9')
        sc(ws, 7, 1, None, al=Alignment(horizontal='left', vertical='center'))
        ws.merge_cells('D5:E7')
        sc(ws, 5, 1, "     Bill  To", fnt=black_font, border=M_LRT)
        sc(ws, 6, 1, "客户名称", fnt=red_font, border=M_ALL)
        sc(ws, 7, 1, "客户地址", fnt=red_font, border=M_ALL)
        sc(ws, 8, 4, "Terms", fnt=black_font, border=M_ALL)
        sc(ws, 8, 5, "Due Date", fnt=black_font, border=M_ALL)
        sc(ws, 9, 4, "付款周期", fnt=red_font, border=M_ALL)
        sc(ws, 9, 5, "到期日", fnt=red_font, border=M_ALL)

        # Column headers
        sc(ws, 11, 1, "Description", fnt=black_font, al=Alignment(horizontal='center'), border=M_ALL)
        ws.merge_cells('A11:B11')
        sc(ws, 11, 3, "Qty", fnt=black_font, border=M_ALL)
        sc(ws, 11, 4, "Rate", fnt=black_font, border=M_ALL)
        sc(ws, 11, 5, "Amount", fnt=black_font, border=M_ALL)
        sc(ws, 11, 2, None, border=M_TB)

        # Line items
        for c in range(1, 6):
            sc(ws, 12, c, None, border=M_LR)
        sc(ws, 12, 1, "行项目.描述.US", fnt=red_font, al=Alignment(horizontal='left', vertical='top'), border=M_LR)
        ws.merge_cells('A12:B12')
        sc(ws, 12, 3, "行项目.数量", fnt=red_font, border=M_LR)
        sc(ws, 12, 4, "行项目.单价", fnt=red_font, border=M_LR)
        sc(ws, 12, 5, "行项目.金额", fnt=red_font, border=M_LR)

        # Remittance + Total
        sc(ws, 14, 1, "Please remit payment in full within 14 days. Thank you for your business!",
           fnt=black_font, al=Alignment(horizontal='left', vertical='center'), border=M_LT)
        ws.merge_cells('A14:C14')
        sc(ws, 14, 2, None, border=M_T)
        sc(ws, 14, 3, None, border=M_T)
        sc(ws, 14, 4, "Total", fnt=black_font, border=M_LTB)
        sc(ws, 14, 5, "合计金额", fnt=red_font, border=M_RTB)

        sc(ws, 15, 1, "发票备注", fnt=red_font, al=Alignment(horizontal='left', vertical='center'), border=M_L)
        ws.merge_cells('A15:C15')
        sc(ws, 15, 4, None, border=M_L)
        sc(ws, 15, 5, None, border=M_R)

        # Bank
        for i, line in enumerate(US_EL_BANK):
            r = 16 + i
            b = M_LB if i == len(US_EL_BANK)-1 else M_L
            sc(ws, r, 1, line, fnt=bank_font, border=b)
            sc(ws, r, 2, None, border=M_B if i == len(US_EL_BANK)-1 else None)
            sc(ws, r, 3, None, border=M_B if i == len(US_EL_BANK)-1 else None)
            sc(ws, r, 4, None, border=M_LB if i == len(US_EL_BANK)-1 else None)
            sc(ws, r, 5, None, border=M_RB if i == len(US_EL_BANK)-1 else None)
        sc(ws, 16, 3, "开票主体", fnt=red_font, al=Alignment(vertical='top'))
        sc(ws, 17, 3, "收款信息", fnt=red_font, al=Alignment(vertical='top'))

def build_ga(ws, company_lines, is_credit=False):
    """US-GA positive or credit note. company_lines=[name, addr1, addr2]."""
    set_cols(ws, {'A': 36, 'B': 14, 'C': 8, 'D': 14, 'E': 16})

    if is_credit:
        sc(ws, 1, 1, company_lines[0], fnt=jp_title,
           al=Alignment(horizontal='left', vertical='center', wrap_text=True))
        for i, line in enumerate(company_lines[1:], 2):
            sc(ws, i, 1, line.strip(), fnt=black_font,
               al=Alignment(horizontal='left', vertical='center', wrap_text=True)) if line.strip() else None
        ws.merge_cells('A1:A3')
        ws.merge_cells('D1:E1')
        sc(ws, 1, 4, "Credit Note", fnt=title_font, al=Alignment(horizontal='right'))
        sc(ws, 2, 4, "Date", fnt=black_font, al=Alignment(horizontal='right'))
        sc(ws, 2, 5, "credit note #", fnt=black_font, al=Alignment(horizontal='right'))
        sc(ws, 3, 4, "开票日期.US", fnt=red_font, al=Alignment(horizontal='right'))
        sc(ws, 3, 5, "发票号码.GA", fnt=red_font, al=Alignment(horizontal='right'))
        ws.merge_cells('D5:E7')
        sc(ws, 5, 1, "     Bill  To", fnt=black_font)
        sc(ws, 6, 1, "客户名称", fnt=red_font)
        sc(ws, 7, 1, "客户地址", fnt=red_font)
        for c, v in {1:"Description", 3:"Qty", 4:"Rate", 5:"Amount"}.items():
            sc(ws, 9, c, v, fnt=black_font, al=Alignment(horizontal='center'))
        ws.merge_cells('A9:B9')
        for c, v in {1:"行项目.描述.US", 3:"行项目.数量", 4:"行项目.单价.NEG", 5:"行项目.金额.NEG"}.items():
            sc(ws, 10, c, v, fnt=red_font, al=Alignment(wrap_text=True, vertical='top'))
        ws.merge_cells('A10:B10')
        sc(ws, 11, 1, "发票备注", fnt=red_font, al=Alignment(horizontal='left', vertical='center'))
        ws.merge_cells('A11:C11')
        sc(ws, 11, 4, "Total", fnt=black_font)
        sc(ws, 11, 5, "合计金额.NEG", fnt=red_font)
        for i, line in enumerate(US_GA_BANK):
            sc(ws, 13+i, 1, line, fnt=bank_font)
        sc(ws, 13, 3, "开票主体", fnt=red_font, al=Alignment(vertical='top'))
        sc(ws, 14, 3, "收款信息", fnt=red_font, al=Alignment(vertical='top'))
    else:
        sc(ws, 1, 1, company_lines[0], fnt=jp_title,
           al=Alignment(horizontal='left', vertical='center', wrap_text=True))
        for i, line in enumerate(company_lines[1:], 2):
            sc(ws, i, 1, line.strip(), fnt=black_font,
               al=Alignment(horizontal='left', vertical='center', wrap_text=True)) if line.strip() else None
        ws.merge_cells('A1:A3')
        ws.merge_cells('B1:D1')
        sc(ws, 1, 5, "Invoice", fnt=title_font, al=Alignment(horizontal='right'))
        sc(ws, 2, 4, "Date", fnt=black_font, al=Alignment(horizontal='right'))
        sc(ws, 2, 5, "Invoice #", fnt=black_font, al=Alignment(horizontal='right'))
        sc(ws, 3, 4, "开票日期.US", fnt=red_font, al=Alignment(horizontal='right'))
        sc(ws, 3, 5, "发票号码.GA", fnt=red_font, al=Alignment(horizontal='right'))
        ws.merge_cells('B5:C9')
        ws.merge_cells('A7:A9')
        sc(ws, 7, 1, None, al=Alignment(horizontal='left', vertical='center'))
        ws.merge_cells('D5:E7')
        sc(ws, 5, 1, "     Bill  To", fnt=black_font, border=M_LRT)
        sc(ws, 6, 1, "客户名称", fnt=red_font, border=M_ALL)
        sc(ws, 7, 1, "客户地址", fnt=red_font, border=M_ALL)
        sc(ws, 8, 4, "Terms", fnt=black_font, border=M_ALL)
        sc(ws, 8, 5, "Due Date", fnt=black_font, border=M_ALL)
        sc(ws, 9, 4, "付款周期", fnt=red_font, border=M_ALL)
        sc(ws, 9, 5, "到期日", fnt=red_font, border=M_ALL)
        sc(ws, 11, 1, "Description", fnt=black_font, border=M_ALL)
        ws.merge_cells('A11:B11')
        sc(ws, 11, 2, None, border=M_TB)
        sc(ws, 11, 3, "Qty", fnt=black_font, border=M_ALL)
        sc(ws, 11, 4, "Rate", fnt=black_font, border=M_ALL)
        sc(ws, 11, 5, "Amount", fnt=black_font, border=M_ALL)
        for c in range(1, 6):
            sc(ws, 12, c, None, border=M_LR)
        sc(ws, 12, 1, "行项目.描述.US", fnt=red_font, al=Alignment(horizontal='left', vertical='top'), border=M_LR)
        ws.merge_cells('A12:B12')
        sc(ws, 12, 3, "行项目.数量", fnt=red_font, border=M_LR)
        sc(ws, 12, 4, "行项目.单价", fnt=red_font, border=M_LR)
        sc(ws, 12, 5, "行项目.金额", fnt=red_font, border=M_LR)
        sc(ws, 14, 1, "Please remit payment in full within 14 days. Thank you for your business!",
           fnt=black_font, al=Alignment(horizontal='left', vertical='center'), border=M_LT)
        ws.merge_cells('A14:C14')
        sc(ws, 14, 2, None, border=M_T)
        sc(ws, 14, 3, None, border=M_T)
        sc(ws, 14, 4, "Total", fnt=black_font, border=M_LTB)
        sc(ws, 14, 5, "合计金额", fnt=red_font, border=M_RTB)
        sc(ws, 15, 1, "发票备注", fnt=red_font, al=Alignment(horizontal='left', vertical='center'), border=M_L)
        ws.merge_cells('A15:C15')
        sc(ws, 15, 4, None, border=M_L)
        sc(ws, 15, 5, None, border=M_R)
        for i, line in enumerate(US_GA_BANK):
            r = 16 + i
            b = M_LB if i == len(US_GA_BANK)-1 else M_L
            sc(ws, r, 1, line, fnt=bank_font, border=b)
            sc(ws, r, 2, None, border=M_B if i == len(US_GA_BANK)-1 else None)
            sc(ws, r, 4, None, border=M_LB if i == len(US_GA_BANK)-1 else None)
            sc(ws, r, 5, None, border=M_RB if i == len(US_GA_BANK)-1 else None)
        sc(ws, 16, 3, "开票主体", fnt=red_font, al=Alignment(vertical='top'))
        sc(ws, 17, 3, "收款信息", fnt=red_font, al=Alignment(vertical='top'))

# ══════════════════════════════════════════════════════════════════════
#  13: Japan
# ══════════════════════════════════════════════════════════════════════
def build_jp(ws):
    set_cols(ws, {'A': 30, 'B': 22, 'C': 18, 'D': 8, 'E': 14, 'F': 18})
    ws.merge_cells('A1:F1')
    sc(ws, 4, 5, "発行日", fnt=black_font, border=T_B)
    sc(ws, 4, 6, "开票日期.JP", fnt=red_font, border=T_B)
    sc(ws, 5, 5, "発行番号", fnt=black_font, border=T_TB)
    sc(ws, 5, 6, "发票号码.JP", fnt=red_font, border=T_TB)
    sc(ws, 6, 5, "登録番号", fnt=black_font, border=T_B)
    sc(ws, 6, 6, "T3010001214611", fnt=black_font, border=T_B)
    sc(ws, 7, 5, "日発  株式会社", fnt=black_font)
    sc(ws, 8, 5, "〒101-0051", fnt=black_font)
    sc(ws, 9, 5, "住所：東京都墨田区横川一丁目9番3号", fnt=black_font)
    ws.merge_cells('A10:B11')
    sc(ws, 10, 1, "客户名称", fnt=red_font,
       al=Alignment(horizontal='center', vertical='center'))
    sc(ws, 12, 1, "下記の通り、ご請求申し上げます。", fnt=black_font)
    sc(ws, 16, 1, "お振込み期日：", fnt=black_font)
    sc(ws, 16, 2, "开票日期.JP", fnt=red_font)
    sc(ws, 17, 1, "ご請求金額 ：", fnt=black_font, border=M_B)
    sc(ws, 17, 2, "合计金额(含税)", fnt=red_font, border=M_B)

    ws.merge_cells('A36:D38')
    sc(ws, 36, 1, None, al=Alignment(horizontal='center', vertical='center'))
    ws.merge_cells('A41:B41')
    ws.merge_cells('A42:B42')
    for r in range(19, 36):
        for c in range(1, 7):
            sc(ws, r, c, None, border=T_ALL)
    headers = {1:"NO.", 2:"業務内容", 3:"期間", 4:"数量", 5:"単価", 6:"金額"}
    for c, v in headers.items():
        sc(ws, 19, c, v, fnt=black_font, al=Alignment(horizontal='center'), border=T_ALL)
    item_fields = {1:"行项目.序号", 2:"行项目.品名", 3:"行项目.账期.JP",
                   4:"行项目.数量", 5:"行项目.单价", 6:"行项目.金额"}
    for c, v in item_fields.items():
        sc(ws, 20, c, v, fnt=red_font, al=Alignment(wrap_text=True, vertical='top'), border=T_ALL)

    sc(ws, 36, 1, "发票备注", fnt=red_font, border=T_ALL)
    for c in range(2, 7):
        sc(ws, 36, c, None, border=T_ALL)
    sc(ws, 36, 5, "小計", fnt=black_font, border=T_ALL)
    sc(ws, 36, 6, "小计金额.JP", fnt=red_font, border=T_ALL)
    sc(ws, 37, 1, None, border=T_ALL)
    for c in range(2, 5):
        sc(ws, 37, c, None, border=T_ALL)
    sc(ws, 37, 5, "消費税", fnt=black_font, border=T_ALL)
    sc(ws, 37, 6, "消费税额.JP", fnt=red_font, border=T_ALL)
    sc(ws, 38, 1, None, border=T_ALL)
    for c in range(2, 5):
        sc(ws, 38, c, None, border=T_ALL)
    sc(ws, 38, 5, "合計（税込み）", fnt=black_font, border=T_ALL)
    sc(ws, 38, 6, "合计金额(含税)", fnt=red_font, border=T_ALL)

    sc(ws, 40, 1, "お振込先", fnt=black_font)
    sc(ws, 41, 1, "[振込先] シティバンク、エヌ・エイ 東京支店", fnt=black_font,
       al=Alignment(horizontal='center'), border=T_LRT)
    sc(ws, 41, 2, None, border=T_RT)
    sc(ws, 42, 1, "当座預金 9904695 NICHIHATU KABUSIKIGAISYA", fnt=black_font,
       al=Alignment(horizontal='center'), border=T_LB)
    sc(ws, 42, 2, None, border=T_RB)
    sc(ws, 45, 1, "*恐れ入りますが、お振込み手数料はご負担願います。", fnt=black_font)

# ══════════════════════════════════════════════════════════════════════
#  14: Australia
# ══════════════════════════════════════════════════════════════════════
def build_au(ws):
    set_cols(ws, {'A': 4, 'B': 28, 'C': 16, 'D': 14, 'E': 16, 'F': 16})
    ws.merge_cells('B1:B2')
    sc(ws, 1, 2, "TAX INVOICE", fnt=title_font,
       al=Alignment(horizontal='center', vertical='center'))
    sc(ws, 2, 5, "Invoice Date", fnt=black_font)
    sc(ws, 2, 6, "开票日期.AU", fnt=red_font)
    sc(ws, 3, 2, "Everfast Pty Ltd", fnt=black_font)
    sc(ws, 4, 2, "42-62 POUND ROAD W, DANDENONG SOUTH VIC 3175", fnt=black_font)
    sc(ws, 3, 5, "Invoice Number", fnt=black_font)
    sc(ws, 4, 5, "发票号码.AU", fnt=red_font)
    sc(ws, 3, 6, "开票日期.AU", fnt=red_font)
    sc(ws, 5, 2, "客户名称", fnt=red_font)
    sc(ws, 6, 2, "客户地址", fnt=red_font)
    sc(ws, 5, 5, "Reference", fnt=black_font)
    sc(ws, 5, 6, None)
    sc(ws, 6, 5, None)
    sc(ws, 7, 5, "ABN", fnt=black_font)
    sc(ws, 7, 6, "75 623 285 825", fnt=black_font)

    for c, v in {2:"Description", 3:"Quantity", 4:"Unit Price", 5:"GST", 6:"Amount AUD"}.items():
        sc(ws, 13, c, v, fnt=black_font, al=Alignment(horizontal='center'), border=M_B)

    for r in range(14, 19):
        for c in range(2, 7):
            sc(ws, r, c, None, border=M_B)
    sc(ws, 14, 2, "行项目.描述.AU", fnt=red_font, border=M_B)
    sc(ws, 14, 3, "行项目.数量", fnt=red_font, border=M_B)
    sc(ws, 14, 4, "行项目.单价", fnt=red_font, border=M_B)
    sc(ws, 14, 5, "GST(10%).AU", fnt=red_font, border=M_B)
    sc(ws, 14, 6, "行项目.金额", fnt=red_font, border=M_B)

    sc(ws, 20, 5, "Subtotal", fnt=black_font)
    sc(ws, 20, 6, "小计金额.AU", fnt=red_font)
    sc(ws, 21, 5, "TOTAL GST 10%", fnt=black_font)
    sc(ws, 21, 6, "合计税额.AU", fnt=red_font, border=M_T)
    sc(ws, 22, 5, "TOTAL AUD", fnt=black_font)
    sc(ws, 22, 6, "合计金额(含税)", fnt=red_font, border=M_T)

    sc(ws, 24, 2, "Due Date:", fnt=black_font)
    sc(ws, 24, 3, "到期日", fnt=red_font)
    sc(ws, 25, 2, "Bank Name: Citibank", fnt=black_font)
    sc(ws, 26, 2, "Account Name: Everfast Pty Ltd", fnt=black_font)
    sc(ws, 27, 2, "BSB: 242-000", fnt=black_font)
    sc(ws, 28, 2, "Account No: 236077004", fnt=black_font)
    sc(ws, 29, 2, "发票备注", fnt=red_font)

    # Payment Advice section
    ws.merge_cells('B32:C32')
    sc(ws, 32, 2, "PAYMENT ADVICE", fnt=jp_title)
    sc(ws, 33, 5, "Customer", fnt=black_font)
    sc(ws, 33, 6, "客户名称", fnt=red_font)
    sc(ws, 34, 2, "To:", fnt=black_font)
    sc(ws, 34, 5, "Invoice Number", fnt=black_font)
    sc(ws, 34, 6, "发票号码.AU", fnt=red_font)
    sc(ws, 35, 5, "Amount Due", fnt=black_font)
    sc(ws, 35, 6, "合计金额(含税)", fnt=red_font)
    ws.merge_cells('B36:B37')
    sc(ws, 36, 2, "42-62 POUND ROAD W,\nDANDENONG SOUTH VIC 3175", fnt=black_font)
    sc(ws, 36, 5, "Due Date", fnt=black_font)
    sc(ws, 36, 6, "到期日", fnt=red_font)
    sc(ws, 37, 5, "Amount Enclosed", fnt=gray_font)

# ══════════════════════════════════════════════════════════════════════
#  15-16: ABU Shenzhen / Hong Kong
# ══════════════════════════════════════════════════════════════════════
ABU_INFO = {
    'SZ': {
        'cn': "深圳市云颂国际货运代理有限公司",
        'en': "ShenZhen YunSong International Freight Forwarding Co. LTD",
        'addr_cn': "前湾一路285号创维海外发展大厦1栋402",
        'addr_en': "Room 402, Building 1, SKYWORTH Overseas Development Mansion, No.285 Qianwan First Road",
        'prefix': 'SZ',
    },
    'HK': {
        'cn': "香港雲頌國際貨運有限公司",
        'en': "HONGKONG YUNFREIGHT LIMITED",
        'addr_cn': "香港湾仔骆克道300号浙江兴业大厦（ZJ 300）12楼A室",
        'addr_en': "FLAT/RM A 12/F ZJ 300, 300 LOCKHART ROAD, WANCHAI, HONG KONG",
        'prefix': 'HK',
    },
}

def build_abu(ws, info):
    set_cols(ws, {'A': 4, 'B': 30, 'C': 24, 'D': 14, 'E': 22})

    # Row 2: spacer
    ws.merge_cells('B2:E2')
    sc(ws, 2, 2, "", fnt=black_font,
       al=Alignment(horizontal='center', vertical='center'))

    # Row 3: company info block (merged B:E, field key labels inline)
    ws.merge_cells('B3:E3')
    sc(ws, 3, 2,
       f"{info['cn']}\n{info['en']}\n{info['addr_cn']}\n{info['addr_en']}\n"
       f"（销售方名称（中文）/销售方名称（英文）/销售方地址（中文）/销售方地址（英文））",
       fnt=black_font,
       al=Alignment(horizontal='center', vertical='center'))

    inv_pfx = info['prefix']

    # Row 4: INVOICE title
    ws.merge_cells('B4:E4')
    sc(ws, 4, 2, "INVOICE", fnt=title_font,
       al=Alignment(horizontal='center', vertical='center'))

    # Row 5: To / Date (B5:C5 merged)
    ws.merge_cells('B5:C5')
    sc(ws, 5, 2, "To：", fnt=black_font, border=M_L)
    sc(ws, 5, 4, "Date:", fnt=black_font, border=T_L)
    sc(ws, 5, 5, "开票日期.ABU", fnt=red_font, border=M_R)

    # Row 6: Address / Invoice No
    ws.merge_cells('B6:C6')
    sc(ws, 6, 2, "Address:", fnt=black_font, border=M_L)
    sc(ws, 6, 4, "Invoice No.", fnt=black_font, border=T_L)
    sc(ws, 6, 5, f"发票号码.{inv_pfx}", fnt=red_font, border=M_R)

    # Row 7: Column headers (C7:D7 merged for Period)
    ws.merge_cells('C7:D7')
    sc(ws, 7, 2, "商品/服务描述 Description", fnt=black_font,
       al=Alignment(horizontal='center', vertical='center'), border=M_ALL)
    sc(ws, 7, 3, "期间 Period", fnt=black_font,
       al=Alignment(horizontal='center', vertical='center'))
    sc(ws, 7, 5, "金额（USD） Amount(USD)", fnt=black_font,
       al=Alignment(horizontal='center', vertical='center'), border=M_ALL)

    # Rows 8-16: Line items (C:D merged per row)
    for r in range(8, 17):
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    # First line item row has field keys
    sc(ws, 8, 2, "客户名称", fnt=red_font)
    sc(ws, 9, 2, "客户地址", fnt=red_font)
    sc(ws, 10, 2, "行项目.品名", fnt=red_font, al=Alignment(wrap_text=True, vertical='top'))
    sc(ws, 10, 3, "行项目.账期.ABU", fnt=red_font, al=Alignment(wrap_text=True, vertical='top'))
    sc(ws, 10, 5, "行项目.金额", fnt=red_font, al=Alignment(wrap_text=True, vertical='top'))
    # Border for line item rows
    for r in range(8, 17):
        sc(ws, r, 2, None, border=Border(left=M, right=T, top=T, bottom=T))
        sc(ws, r, 3, None)
        sc(ws, r, 5, None, border=Border(left=T, right=M, top=T, bottom=T))
    # Re-apply cell values after borders
    sc(ws, 8, 2, "客户名称", fnt=red_font)
    sc(ws, 9, 2, "客户地址", fnt=red_font)
    sc(ws, 10, 2, "行项目.品名", fnt=red_font, al=Alignment(wrap_text=True, vertical='top'))
    sc(ws, 10, 3, "行项目.账期.ABU", fnt=red_font, al=Alignment(wrap_text=True, vertical='top'))
    sc(ws, 10, 5, "行项目.金额", fnt=red_font, al=Alignment(wrap_text=True, vertical='top'))

    # Row 17: Total (B17:D17 merged)
    ws.merge_cells('B17:D17')
    sc(ws, 17, 2, "合计（USD）：Total（USD）", fnt=black_font,
       border=Border(left=M, right=T, top=M, bottom=M))
    sc(ws, 17, 5, "合计金额", fnt=red_font, border=M_R)

    # Row 18: Bank header (B18:E18 merged)
    ws.merge_cells('B18:E18')
    sc(ws, 18, 2, "请支付到以下账户（Please Pay To The Following Account）",
       fnt=black_font, al=Alignment(horizontal='center', vertical='center'), border=M_LR)

    # Rows 19-25: Bank detail (C:E merged per row)
    bank_fields = [
        (19, "收款人名称：(Name Of Beneficiary)", "收款人名称"),
        (20, "银行名称 (Bank Name)", "银行名称"),
        (21, "账户名称 Account Name", "账户名称"),
        (22, "账户号 (Account Number)", "账号"),
        (23, "银行代码 (Bank code)", "银行代码"),
        (24, "银行联行码 (Bank Swift Code)", "Swift Code"),
        (25, "银行地址 (Bank Address)", "银行地址"),
    ]
    for r, label, key in bank_fields:
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        sc(ws, r, 2, label, fnt=black_font, border=Border(left=M, right=T, top=T, bottom=T))
        sc(ws, r, 3, key, fnt=red_font, border=Border(left=T, right=M, top=T, bottom=T))

    # Row 26: 发票备注 (C26:E26 merged)
    ws.merge_cells('C26:E26')
    sc(ws, 26, 2, "发票备注：(Comment)", fnt=black_font,
       border=Border(left=M, right=T, top=T, bottom=T))
    sc(ws, 26, 3, "发票备注", fnt=red_font,
       border=Border(left=T, right=M, top=T, bottom=T))
    # Bottom borders
    sc(ws, 26, 2, None, border=M_LB)
    sc(ws, 26, 5, None, border=M_RB)

# ══════════════════════════════════════════════════════════════════════
#  Support sheets
# ══════════════════════════════════════════════════════════════════════
def build_rules(ws):
    set_cols(ws, {'A': 4, 'B': 24, 'C': 30, 'D': 16, 'E': 50, 'F': 40})
    for c, v in {3:"字段key（红字内容）", 4:"取值类型", 5:"取值逻辑", 6:"说明"}.items():
        sc(ws, 4, c, v, fnt=Font(name="Arial", size=10, bold=True, color="FFFFFF"),
           fill=header_fill, border=thin_border, al=Alignment(horizontal='center'))
    def sec(r, col, text):
        sc(ws, r, col, text, fnt=Font(name="Arial", size=10, bold=True, color="4472C4"),
           fill=sub_fill, border=thin_border)
        for cc in range(3, 7):
            sc(ws, r, cc, None, fill=sub_fill, border=thin_border)
    def fl(r, k, t, l, n=""):
        sc(ws, r, 3, k, fnt=red_font, border=thin_border)
        sc(ws, r, 4, t, fnt=black_font, border=thin_border)
        sc(ws, r, 5, l, fnt=gray_font, border=thin_border, al=Alignment(wrap_text=True))
        sc(ws, r, 6, n, fnt=gray_font, border=thin_border, al=Alignment(wrap_text=True))
    r = 5
    sec(r, 2, "申请单映射字段"); r += 1
    for v in [("客户名称","字段映射","申请单->客户名称","所有模板通用"),
              ("客户地址","字段映射","申请单->客户地址",""),
              ("开票日期.US","规则计算","ITS->账期最后一天/审核当天","格式:mm/dd/yyyy - EL/GA/CA"),
              ("开票日期.VC","规则计算","ITS->账期最后一天/审核当天","格式:Mon/dd/yyyy - 温哥华"),
              ("开票日期.JP","规则计算","ITS->账期最后一天/审核当天","格式:yyyy年mm月dd日"),
              ("开票日期.AU","规则计算","ITS->账期最后一天/审核当天","格式:dd-mm-yyyy"),
              ("开票日期.ABU","规则计算","审核当天","格式:yyyy-mm-dd - ABU"),
              ("发票号码.EL","规则计算","公司简称+US+yyyymmdd+B+4seq","美国EL"),
              ("发票号码.GA","规则计算","公司简称+US+yyyymmdd+B+4seq","美国GA"),
              ("发票号码.JP","规则计算","NF+yyyy-mmdd+2seq","日本"),
              ("发票号码.AU","规则计算","EPAUS+yyyymm+B+3seq","澳洲"),
              ("发票号码.CA","规则计算","GCCA+yyyymm+B+4seq","加拿大"),
              ("发票号码.VC","规则计算","YTTY+yyyymm+B+4seq","温哥华"),
              ("发票号码.SZ","规则计算","YSSZ+yyyymm+F+4seq","深圳"),
              ("发票号码.HK","规则计算","YSHK+yyyymm+F+4seq","香港"),
              ("付款周期","字段映射","申请单->付款条款","如Net 14, Net 30"),
              ("到期日","公式计算","开票日期+付款周期天数","EL/GA/加拿大/温哥华")]:
        fl(r, *v); r += 1
    sec(r, 2, "系统生成/计算字段"); r += 1
    for v in [("币种","字段映射","申请单->币种",""),
              ("发票备注","字段映射","申请单->发票备注","无值则留空"),
              ("合计金额","公式计算","SUM(行项目.金额)","不含税合计"),
              ("合计金额.NEG","公式计算","SUM(行项目.金额.NEG)","负数模版"),
              ("合计税额","公式计算","SUM(行项目.税额)","CA/VC"),
              ("合计税额.NEG","公式计算","SUM(行项目.税额.NEG)","负数模版"),
              ("合计含税金额","公式计算","SUM(行项目.含税金额)","CA/VC"),
              ("合计含税金额.NEG","公式计算","SUM(行项目.含税金额.NEG)","负数模版"),
              ("小计金额.JP","公式计算","SUM(行项目.金额)","日本谷仓使用"),
              ("消费税额.JP","公式计算","小计金额*10%","日本谷仓消费税"),
              ("小计金额.AU","公式计算","SUM(行项目.金额)","澳洲使用"),
              ("合计税额.AU","公式计算","小计金额*10%","澳洲GST总额"),
              ("合计含税金额","公式计算","合计金额+合计税额","CA/VC/JP/AU")]:
        fl(r, *v); r += 1
    sec(r, 2, "行项目区字段"); r += 1
    for v in [("行项目.描述.US","字段映射","行项目->品名+' for '+账期(m/d/y-m/d/y)","EL/GA/CA/VC"),
              ("行项目.描述.AU","字段映射","行项目->品名+'--'+账期(dd/mm/yyyy-dd/mm/yyyy)","澳洲"),
              ("行项目.品名","字段映射","行项目->品名","日本/ABU"),
              ("行项目.序号","字段映射","行项目->序号","日本"),
              ("行项目.账期.US","字段映射","行项目->账期","格式:mm/dd/yyyy-mm/dd/yyyy"),
              ("行项目.账期.JP","字段映射","行项目->账期","格式:yyyy/mm/dd-yyyy/mm/dd"),
              ("行项目.账期.AU","字段映射","行项目->账期","格式:dd/mm/yyyy-dd/mm/yyyy"),
              ("行项目.账期.ABU","字段映射","行项目->账期","格式:yyyy-mm-dd至yyyy-mm-dd"),
              ("行项目.数量","字段映射","行项目->数量","默认1"),
              ("行项目.单价","字段映射","行项目->单价",""),
              ("行项目.金额","字段映射","行项目->金额","不含税"),
              ("行项目.金额.NEG","公式计算","行项目->金额 * (-1)","负数模版取反"),
              ("行项目.单价.NEG","公式计算","行项目->单价 * (-1)","负数模版取反"),
              ("行项目.税额","公式计算","金额 * 税率","CA/VC"),
              ("行项目.税额.NEG","公式计算","金额 * 税率 * (-1)","负数模版取反"),
              ("行项目.含税金额","公式计算","金额 + 税额","CA/VC"),
              ("行项目.含税金额.NEG","公式计算","(金额+税额) * (-1)","负数模版取反"),
              ("GST(10%).AU","公式计算","金额 * 10%","澳洲")]:
        fl(r, *v); r += 1
    sec(r, 2, "主体预设字段"); r += 1
    for v in [("收款人名称","字段映射","主体->收款人名称","ABU"),
              ("银行名称","字段映射","主体->银行名称","ABU"),
              ("账户名称","字段映射","主体->账户名称","ABU"),
              ("账号","字段映射","主体->账号","ABU"),
              ("银行代码","字段映射","主体->银行代码","ABU"),
              ("Swift Code","字段映射","主体->Swift Code","ABU"),
              ("银行地址","字段映射","主体->银行地址","ABU")]:
        fl(r, *v); r += 1
    sec(r, 2, "开票主体/收款信息"); r += 1
    for v in [("开票主体","主体预设","主体信息->名称","FBU"),
              ("收款信息","主体预设","主体信息->收款银行信息","FBU"),
              ("销售方名称（中文）","主体预设","主体->中文名称","ABU"),
              ("销售方名称（英文）","主体预设","主体->英文名称","ABU"),
              ("销售方地址（中文）","主体预设","主体->中文地址","ABU"),
              ("销售方地址（英文）","主体预设","主体->英文地址","ABU")]:
        fl(r, *v); r += 1
    r += 1
    sc(ws, r, 2, "===== 字段取值差异说明 =====",
       fnt=Font(name="Arial", size=10, bold=True, color="4472C4")); r += 1
    notes = [
        "【发票号码】各主体编码规则不同，详见[模版总览]Sheet及[字段取值规则]中对应字段。已拆分发票号码.EL/.GA/.JP/.AU/.CA/.VC/.SZ/.HK",
        "【开票日期】日期格式不同：US(m/d/y), VC(Mon/d/y), JP(y年m月d日), AU(d-m-y), ABU(y-m-d)。已拆分开票日期.US/.VC/.JP/.AU/.ABU",
        "【行项目.描述】拼接规则不同：US国家(EL/GA/CA/VC)=品名+for+账期，AU=品名+--+账期，JP/ABU=品名(单独列)。已拆分行项目.描述.US/.AU，JP/ABU用行项目.品名",
        "【行项目.账期】日期格式不同：US=m/d/y-m/d/y，JP=y/m/d-y/m/d，AU=d/m/y-d/m/y，ABU=y-m-d至y-m-d。已拆分行项目.账期.US/.JP/.AU/.ABU",
        "【行项目.税额/含税金额】仅加拿大和温哥华使用",
        "【GST(10%).AU】仅澳洲使用",
        "【小计金额.JP/消费税额.JP】仅日本使用",
        "【小计金额.AU/合计税额.AU】仅澳洲使用",
        "【付款周期】所有模板通用，部分模板固定值则黑字",
        "【.NEG后缀】负数模版中金额/税额/含税金额字段加.NEG后缀，系统取申请金额的相反数",
    ]
    for n in notes:
        sc(ws, r, 2, n, fnt=gray_font, al=Alignment(wrap_text=True)); r += 1

def build_overview(ws):
    set_cols(ws, {'A': 12, 'B': 24, 'C': 10, 'D': 18, 'E': 10, 'F': 10, 'G': 30})
    sc(ws, 1, 2, "模版总览",
       fnt=Font(name="Arial", size=14, bold=True, color="4472C4"))
    sc(ws, 2, 2, "所有发票模版统一注册管理，系统通过 模版编码(template_id) 唯一识别每条模版，导入更新时以此匹配",
       fnt=gray_font)

    headers = {2:"模版编码\nTemplate ID", 3:"Sheet名称", 4:"主体",
               5:"税率", 6:"正/负数", 7:"备注"}
    for c, v in headers.items():
        sc(ws, 4, c, v, fnt=Font(name="Arial", size=9, bold=True, color="FFFFFF"),
           fill=header_fill, border=thin_border, al=Alignment(horizontal='center', wrap_text=True))

    rules = [
        ("CA-GST5",   "FBU-加拿大-5%GST",    "加拿大", "GST 5%",  "正数", "正负数同B"),
        ("CA-GST5N",  "FBU-加拿大-5%GST-负数","加拿大","GST 5%",  "负数", "正负数同B"),
        ("CA-HST13",  "FBU-加拿大-13%HST",   "加拿大", "HST 13%", "正数", "正负数同B"),
        ("CA-HST13N", "FBU-加拿大-13%HST-负数","加拿大","HST 13%","负数", "正负数同B"),
        ("VC-GST",    "FBU-温哥华-GST",       "温哥华", "GST 0%",  "正数", ""),
        ("VC-GSTN",   "FBU-温哥华-GST-负数",  "温哥华", "GST 0%",  "负数", ""),
        ("VC-HST",    "FBU-温哥华-HST",       "温哥华", "HST 13%", "正数", ""),
        ("VC-HSTN",   "FBU-温哥华-HST-负数",  "温哥华", "HST 13%", "负数", ""),
        ("EL",        "FBU-美国EL",           "美国EL", "不适用", "正数", "正负数同B"),
        ("EL-N",      "FBU-美国EL-负数",      "美国EL", "不适用", "负数", "正负数同B"),
        ("GA",        "FBU-美国GA",           "美国GA", "不适用", "正数", "正负数同B"),
        ("GA-N",      "FBU-美国GA-负数",      "美国GA", "不适用", "负数", "正负数同B"),
        ("JP",        "FBU-日本谷仓",         "日本",   "不适用", "不区分", ""),
        ("AU",        "FBU-澳洲谷仓",         "澳洲",   "不适用", "不区分", ""),
        ("ABU-SZ",    "ABU-深圳云颂",         "深圳",   "不适用", "不区分", ""),
        ("ABU-HK",    "ABU-香港云颂",         "香港",   "不适用", "不区分", ""),
    ]
    for i, (tid, sheet, entity, tax, sign, note) in enumerate(rules):
        r = 5 + i
        sc(ws, r, 2, tid, fnt=Font(name="Consolas", size=9, bold=True, color="4472C4"),
           border=thin_border, al=Alignment(horizontal='center'))
        sc(ws, r, 3, sheet, fnt=black_font, border=thin_border)
        sc(ws, r, 4, entity, fnt=black_font, border=thin_border)
        sc(ws, r, 5, tax, fnt=black_font, border=thin_border, al=Alignment(horizontal='center'))
        sc(ws, r, 6, sign, fnt=black_font, border=thin_border, al=Alignment(horizontal='center'))
        sc(ws, r, 7, note, fnt=gray_font, border=thin_border)

    # Footer notes
    rn = 6 + len(rules) + 1
    sc(ws, rn, 2, "系统操作说明", fnt=Font(name="Arial", size=10, bold=True, color="4472C4")); rn += 1
    notes = [
        "【导入新增】上传xlsx → 系统解析模版内容，按模版编码(template_id)查找是否已存在 → 不存在则创建新模版记录",
        "【更新导入】上传xlsx → 系统按模版编码匹配已存在模版 → 版本号+1，替换模版定义 → 旧版归档，历史发票追溯",
        "【禁用】模版状态改为禁用 → 新开票申请不可选此模版 → 已审批通过的发票不受影响",
        "【开票选版】开票申请单审批完成后，进入开票环节时系统自动使用这类模版的最新版本",
        "【匹配规则】系统根据 开票主体 + 税率 + 金额正负 自动匹配对应的模版编码，在注册表中查找最新版本",
    ]
    for n in notes:
        sc(ws, rn, 2, n, fnt=gray_font, al=Alignment(wrap_text=True)); rn += 1

def build_guide(ws):
    set_cols(ws, {'A': 30, 'B': 66})
    sc(ws, 1, 1, "发票模板导入规则说明", fnt=Font(name="Arial", size=14, bold=True, color="000000"))
    sections = [
        ("一、核心颜色规则", [
            ("黑色字体", "静态文本，系统原样输出"),
            ("红色字体(RGB#FF0000)", "字段key，系统在[字段取值规则]查找取值逻辑"),
        ]),
        ("二、行项目区", [
            ("行项目字段", "含'行项目.'前缀的字段所在行为列模板，按数据行数重复渲染"),
        ]),
        ("三、字段作用域", [
            ("同名不同规", "同一字段在不同模板取值规则不同时，用后缀区分（如开票日期.US、开票日期.JP）"),
            ("共享复用", "相同后缀的字段共用同一取值规则，减少重复配置"),
        ]),
        ("四、正数/负数发票", [
            ("双模板并存", "加拿大/温哥华/美国EL/美国GA各维护正数和负数两个模板"),
            ("自动选择", "金额≥0用正数，<0用负数"),
        ]),
        ("五、多税率模板", [
            ("分税率模板", "加拿大/温哥华按税率分组（GST/HST）各维护独立模板"),
            ("自动匹配", "系统根据开票税率选择对应模板"),
        ]),
        ("六、关键约束", [
            ("银行区", "FBU: 黑字固定; ABU: 红字从主体预设映射"),
            ("编码规则分离", "发票号码拼接规则在[号码编码规则]Sheet管理"),
            ("字段取值规则", "所有字段key的取值逻辑集中在[字段取值规则]Sheet"),
        ]),
        ("七、模版生命周期管理", [
            ("模版编码", "每个模版有唯一编码(template_id)，集中在[号码编码规则]维护，导入更新时以此匹配"),
            ("三区自动划分", "系统自动识别: 行项目以上=Header, 行项目行=数据模版, 行项目以下=Footer(动态下移)"),
            ("新增导入", "上传xlsx→系统按template_id查找→不存在则创建新模版"),
            ("更新导入", "上传xlsx→匹配已存在模版→版本+1替换→旧版归档(历史发票可追溯)"),
            ("禁用", "模版禁用后新申请不可选，已审批发票不受影响"),
            ("开票用最新版", "审批完成后进入开票环节，系统自动使用该模版最新版本"),
        ]),
    ]
    r = 3
    for title, items in sections:
        sc(ws, r, 1, title, fnt=Font(name="Arial", size=11, bold=True, color="4472C4")); r += 1
        for label, detail in items:
            sc(ws, r, 1, label, fnt=black_font, al=Alignment(wrap_text=True))
            sc(ws, r, 2, detail, fnt=gray_font, al=Alignment(wrap_text=True)); r += 1
        r += 1

# ══════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Helper to create template sheets
    def add(name, fn):
        ws = wb.create_sheet(title=name)
        fn(ws)

    # 1-4: Canada (4 templates)
    add("FBU-加拿大-5%GST", lambda ws: build_ca(ws, 'GST 5%', CA_TAX_INFO['GST5'][1]))
    add("FBU-加拿大-5%GST-负数", lambda ws: build_ca(ws, 'GST 5%', CA_TAX_INFO['GST5'][1], is_credit=True))
    add("FBU-加拿大-13%HST", lambda ws: build_ca(ws, 'HST 13%', CA_TAX_INFO['HST13'][1]))
    add("FBU-加拿大-13%HST-负数", lambda ws: build_ca(ws, 'HST 13%', CA_TAX_INFO['HST13'][1], is_credit=True))

    # 5-8: Vancouver (4 templates)
    add("FBU-温哥华-GST", lambda ws: build_vc(ws, 'GST 0%', VC_TAX_INFO['GST'][1]))
    add("FBU-温哥华-GST-负数", lambda ws: build_vc(ws, 'GST 0%', VC_TAX_INFO['GST'][1], is_credit=True))
    add("FBU-温哥华-HST", lambda ws: build_vc(ws, 'HST 13%', VC_TAX_INFO['HST'][1]))
    add("FBU-温哥华-HST-负数", lambda ws: build_vc(ws, 'HST 13%', VC_TAX_INFO['HST'][1], is_credit=True))

    # 9-10: US-EL
    el_lines = ["US ELOGISTICS SERVICE CORP", "1100 CRANBURY SOUTH RIVER RD", "MONROE, NJ 08831"]
    add("FBU-美国EL", lambda ws: build_el(ws, el_lines))
    add("FBU-美国EL-负数", lambda ws: build_el(ws, el_lines, is_credit=True))

    # 11-12: US-GA
    ga_lines = ["ELOGISTICS (GA) SERVICE CORP", "240 THE BLUFFS AUSTELL, GA 30168", ""]
    add("FBU-美国GA", lambda ws: build_ga(ws, ga_lines))
    add("FBU-美国GA-负数", lambda ws: build_ga(ws, ga_lines, is_credit=True))

    # 13: Japan
    add("FBU-日本谷仓", build_jp)

    # 14: Australia
    add("FBU-澳洲谷仓", build_au)

    # 15-16: ABU
    add("ABU-深圳云颂", lambda ws: build_abu(ws, ABU_INFO['SZ']))
    add("ABU-香港云颂", lambda ws: build_abu(ws, ABU_INFO['HK']))

    # Support sheets (模版总览 first)
    for name, fn in [("模版总览", build_overview), ("字段取值规则", build_rules), ("导入规则说明", build_guide)]:
        ws = wb.create_sheet(title=name, index=0 if name == "模版总览" else None)
        fn(ws)

    out = "/Users/a/Desktop/ZT/FSSC/invoice/发票模板导入格式示例.xlsx"
    wb.save(out)
    print(f"Saved: {out}")
    print(f"Sheets ({len(wb.sheetnames)}): {wb.sheetnames}")

if __name__ == "__main__":
    main()
